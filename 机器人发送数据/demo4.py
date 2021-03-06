import win32api
import win32con
import win32gui
import pandas as pd
import time
import win32clipboard as w
import requests
import json
from win32com.client import Dispatch, DispatchEx
from PIL import ImageGrab, Image
import uuid
import xlwt
import openpyxl
import openpyxl.styles
from openpyxl.styles import PatternFill


session = requests.Session()
first_url_day = 'http://1.192.88.18:8115/hnAqi/v1.0/api/air/dayAqi2018_county'
first_url_month = 'http://1.192.88.18:8115/hnAqi/v1.0/api/air/airreport2018_county'
first_url_year = 'http://1.192.88.18:8115/hnAqi/v1.0/api/air/airreport2018_county '
url_list = 'http://1.192.88.18:8115/hnAqi/v1.0/api/air/getCityStationDetail'
headers = {
    'User-Agent':'Dalvik/2.1.0 (Linux; U; Android 7.1.2; LIO-AN00 Build/N2G48H)'
}


def FindWindow(chatroom):
    win = win32gui.FindWindow(None, chatroom)
    print("找到群聊窗口：%x" % win)
    if win != 0:
        win32gui.ShowWindow(win, win32con.SW_SHOWMINIMIZED)
        win32gui.ShowWindow(win, win32con.SW_SHOWNORMAL)
        win32gui.ShowWindow(win, win32con.SW_SHOW)
        win32gui.SetWindowPos(win, win32con.HWND_TOPMOST, 100, 100, 300, 300, win32con.SWP_SHOWWINDOW)
        win32gui.SetForegroundWindow(win)  # 获取控制
        time.sleep(1)
    else:
        print('请注意：找不到【%s】这个人（或群），请激活窗口！' % chatroom)
        exit()
def CloseWindow(chatroom):
    win = win32gui.FindWindow(None, chatroom)
    #print("找到关闭窗口：%x" % win)
    time.sleep(3)
    win32gui.ShowWindow(win, win32con.SW_SHOWMINIMIZED)

def setText(aString):
    w.OpenClipboard()
    w.EmptyClipboard()
    w.SetClipboardData(win32con.CF_UNICODETEXT, aString)
    w.CloseClipboard()


def ctrlV():
    win32api.keybd_event(17, 0, 0, 0)  # ctrl键位码是17
    win32api.keybd_event(86, 0, 0, 0)  # v键位码是86
    win32api.keybd_event(86, 0, win32con.KEYEVENTF_KEYUP, 0)  # 释放按键
    win32api.keybd_event(17, 0, win32con.KEYEVENTF_KEYUP, 0)


def altS():
    win32api.keybd_event(18, 0, 0, 0)  # Alt键位码
    win32api.keybd_event(83, 0, 0, 0)  # s键位码
    win32api.keybd_event(18, 0, win32con.KEYEVENTF_KEYUP, 0)  # 释放按键
    win32api.keybd_event(83, 0, win32con.KEYEVENTF_KEYUP, 0)


def sendText(chatrooms,text):
    for chatroom in chatrooms:
        FindWindow(chatroom)
        #文字首行留空，防止带表情复制不完全
        setText(" "+text)
        time.sleep(1)
        ctrlV()
        time.sleep(1)
        altS()
        CloseWindow(chatroom)


def year(url):
    data = {
        'end':'2020-10-22',
        'sort':'asc',
        'start':'2020-01-01'
    }
    response = session.post(url= url,data=data,headers = headers).text
    print(response)
    return response

def month(url):
    data = {
        'sort':'asc',
        'month':'2020-07'
    }
    response = session.post(url= url,data=data,headers = headers).text
    return response

def day(url):
    data = {
        'sort':'asc',
        'time':'2020-09-27'
    }
    response = session.post(url= url,data=data,headers = headers).text
    return response


def real(url):
    response = session.get(url= url_list,headers = headers).text
    return response

def save_excel():
    book = xlwt.Workbook()
    sheet = book.add_sheet('淮阳县空气质量数据')
    n = 1
    sheet.write(0,0,'区县')
    sheet.write(0,1,'时间')
    sheet.write(0,2,'CO')
    sheet.write(0,3,'O3')
    sheet.write(0,4,'SO2')
    sheet.write(0,5,'NO2')
    sheet.write(0,6,'PM10')
    sheet.write(0,7,'PM2.5')
    sheet.write(0,8,'AQI')
    sheet.write(0,9,'首要污染物')
    time.sleep(5)
    l = real(url_list)
    data = json.loads(l)['detail']
    for k in data:
        if k['CITY'] in ['沈丘县','商水县','西华县','扶沟县','郸城县','淮阳县','太康县','鹿邑县','项城市']:
            sheet.write(n, 0, k['CITY'])
            try:
                sheet.write(n, 1, k["MONIDATE"].split(" ")[1])
            except:
                sheet.write(n, 1, "无")
            sheet.write(n, 2, k['CO'])
            sheet.write(n, 3, k['O3'])
            sheet.write(n, 4, k['SO2'])
            sheet.write(n, 5, k['NO2'])
            sheet.write(n, 6, k['PM10'])
            sheet.write(n, 7, k['PM25'])
            sheet.write(n, 8, k['AQI'])
            sheet.write(n, 9, k['PRIMARYPOLLUTANT'])
            n+=1
    book.save('周口市区县数据.xls')

def excel_catch_screen():
    """ 对excel的表格区域进行截图——用例：excel_catch_screen(r"E:\年周口市区县27日数据.xls", "淮阳县空气质量数据", "A1:J10")"""
    # pythoncom.CoInitialize()  # excel多线程相关

    excel = DispatchEx("Excel.Application")  # 启动excel
    excel.Visible = True  # 可视化
    excel.DisplayAlerts = False  # 是否显示警告
    wb = excel.Workbooks.Open(r"D:\Program Files\pycharm\机器人发送数据\周口市区县数据排名充填.xlsx")  # 打开excel
    ws = wb.Sheets("Sheet1")  # 选择sheet
    ws.Range("A1:L10").CopyPicture()  # 复制图片区域
    ws.Paste()  # 粘贴 ws.Paste(ws.Range('B1'))  # 将图片移动到具体位置
    name = str(uuid.uuid4())  # 重命名唯一值
    new_shape_name = name[:6]
    excel.Selection.ShapeRange.Name = new_shape_name  # 将刚刚选择的Shape重命名，避免与已有图片混淆

    ws.Shapes(new_shape_name).Copy()  # 选择图片
    img = ImageGrab.grabclipboard()  # 获取剪贴板的图片数据
    # if not img_name:
    #     img_name = name + ".PNG"
    # img.save(img_name)  # 保存图片
    wb.Close(SaveChanges=0)  # 关闭工作薄，不保存
    excel.Quit()  # 退出excel
    # pythoncom.CoUninitialize()

def excel_rank():
    df = pd.read_excel(r'D:\Program Files\pycharm\机器人发送数据\周口市区县数据.xls')
    df['PM2.5排名'] = df['PM2.5'].rank(method='min',ascending=True)
    df['PM10排名'] = df['PM10'].rank(method='min',ascending=True)
    df = df.sort_values(by='PM2.5排名')
    df.reset_index(drop=True, inplace=True)
    df.to_excel(r"D:\Program Files\pycharm\机器人发送数据\周口市区县数据排名.xlsx",index=False)

def excel_c():
    wb = openpyxl.load_workbook(r"D:\Program Files\pycharm\机器人发送数据\周口市区县数据排名.xlsx")
    sheet = wb["Sheet1"]
    n = 0
    fille=PatternFill("solid",fgColor="FFBB02")
    for i in range(1,11):
        if sheet.cell(i, 1).value == "淮阳县":
            n = i
    for j in range(1,13):
        sheet.cell(n, j).fill = fille
    pm25time = sheet.cell(n, 2).value
    pm25nong = sheet.cell(n, 8).value
    pm10nong = sheet.cell(n, 7).value
    pm25rank = sheet.cell(n, 11).value
    pm10rank = sheet.cell(n, 12).value
    wb.save(r"D:\Program Files\pycharm\机器人发送数据\周口市区县数据排名充填.xlsx")
    return (pm25time,pm25nong,pm25rank,pm10nong,pm10rank)



if __name__ == '__main__':
    while True:
        try:
            FindWindow("淮阳区环境攻坚群")
            CloseWindow("淮阳区环境攻坚群")
            save_excel()
            print("已获取完数据")
            excel_rank()
            print("已对数据排名")
            l = excel_c()
            print(l)
            print("已对数据充填")
            excel_catch_screen()
            print("已截图")
            ctrlV()
            time.sleep(1)
            altS()
            time.sleep(1)
            print("已发送截图")
            time.sleep(5)
            FindWindow("淮阳区环境攻坚群")
            CloseWindow("淮阳区环境攻坚群")
            setText("【实时播报】：\n淮阳县{}时，PM2.5浓度为{}μg/m3，在全市9个区县中排名第{}；PM10浓度为{}μg/m3，在全市9个区县中排名第{}。数据来自《河南省空气质量实况与播报app》。".format(l[0],l[1],l[2],l[3],l[4]))
            time.sleep(1)
            ctrlV()
            time.sleep(1)
            altS()
            time.sleep(1)
            print("已发送app")
            time.sleep(3590)
        except:
            time.sleep(3590)
            pass